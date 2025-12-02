from __future__ import annotations

import calendar
import csv
import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict

import pandas as pd

from .paths import ensure_app_data_dir

MONTH_CHOICES = [
    (1, "Janeiro"),
    (2, "Fevereiro"),
    (3, "Março"),
    (4, "Abril"),
    (5, "Maio"),
    (6, "Junho"),
    (7, "Julho"),
    (8, "Agosto"),
    (9, "Setembro"),
    (10, "Outubro"),
    (11, "Novembro"),
    (12, "Dezembro"),
]

MONTH_NAME_TO_NUMBER = {name: number for number, name in MONTH_CHOICES}
MONTH_NUMBER_TO_NAME = {number: name for number, name in MONTH_CHOICES}
WEEKDAY_LABELS = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


def get_weekday_labels(first_weekday: int = 0) -> list[str]:
    """Retorna labels dos dias da semana reordenados conforme primeiro dia configurado."""
    return WEEKDAY_LABELS[first_weekday:] + WEEKDAY_LABELS[:first_weekday]


def _parse_calendar_cell(cell: str | None) -> tuple[int | None, int | None]:
    text = (cell or "").strip()
    if not text:
        return None, None

    if text.endswith(")") and "(" in text:
        day_part, value_part = text[:-1].split("(", 1)
        return _safe_int(day_part), _safe_int(value_part)

    for separator in ("->", ":", "-", "|"):
        if separator in text:
            day_part, value_part = text.split(separator, 1)
            return _safe_int(day_part), _safe_int(value_part)

    return _safe_int(text), None


def _safe_int(value) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _month_key(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def _get_custom_week_start(target_date: date, first_weekday: int) -> date:
    """
    Calcula o início da semana baseado na configuração personalizada do usuário.
    
    Args:
        target_date: Data de referência
        first_weekday: Primeiro dia da semana (0=Monday, 1=Tuesday, ..., 6=Sunday)
    
    Returns:
        Data do início da semana
    """
    current_weekday = target_date.weekday()  # 0=Monday, 6=Sunday
    days_from_week_start = (current_weekday - first_weekday) % 7
    return target_date - timedelta(days=days_from_week_start)


def _get_custom_week_number(target_date: date, first_weekday: int) -> tuple[int, int]:
    """
    Calcula ano e número da semana baseado na configuração personalizada do usuário.
    
    Args:
        target_date: Data de referência
        first_weekday: Primeiro dia da semana (0=Monday, 1=Tuesday, ..., 6=Sunday)
    
    Returns:
        Tupla (ano, numero_da_semana) baseada no calendário personalizado
    """
    # Pega o início do ano
    year_start = date(target_date.year, 1, 1)
    
    # Encontra o primeiro dia da primeira semana do ano
    first_week_start = _get_custom_week_start(year_start, first_weekday)
    
    # Se a primeira semana começou no ano anterior, ajusta
    if first_week_start.year < target_date.year:
        first_week_start = _get_custom_week_start(date(target_date.year, 1, 7), first_weekday)
    
    # Calcula início da semana para a data alvo
    week_start = _get_custom_week_start(target_date, first_weekday)
    
    # Se a semana começou no ano anterior, pertence ao ano anterior
    if week_start.year < target_date.year:
        # Última semana do ano anterior
        prev_year_end = date(target_date.year - 1, 12, 31)
        return _get_custom_week_number(prev_year_end, first_weekday)
    
    # Calcula número da semana
    days_diff = (week_start - first_week_start).days
    week_number = (days_diff // 7) + 1
    
    return target_date.year, week_number


@dataclass
class WeeklyTarget:
    """Representa a meta semanal configurada pelo usuário."""
    year: int
    month: int  # Mês (1-12)
    week_in_month: int  # Número da semana no mês (1-5)
    expected: int = 0

    def week_key(self) -> str:
        """Chave única para a semana no formato ano-mês-semana."""
        return f"{self.year}-{self.month:02d}-S{self.week_in_month}"


@dataclass
class WeekConfig:
    """
    Configuração personalizada de uma semana específica.
    
    Permite ao usuário definir quando sua semana de trabalho começa
    e quantos dias ela terá.
    """
    year: int
    month: int
    week_in_month: int  # Número da semana no mês (1-5)
    start_date: str  # Data de início no formato YYYY-MM-DD
    work_days: int = 6  # Quantidade de dias de trabalho (padrão 6)
    first_weekday: int = 0  # Primeiro dia da semana (0=Seg, 6=Dom)
    
    def config_key(self) -> str:
        """Chave única para a configuração da semana."""
        return f"{self.year}-{self.month:02d}-S{self.week_in_month}"
    
    def get_start_date(self) -> date:
        """Retorna a data de início como objeto date."""
        return datetime.strptime(self.start_date, "%Y-%m-%d").date()
    
    def get_end_date(self) -> date:
        """Retorna a data de fim baseada nos dias de trabalho."""
        start = self.get_start_date()
        return start + timedelta(days=self.work_days - 1)
    
    def get_all_dates(self) -> list[date]:
        """Retorna lista de todas as datas da semana de trabalho."""
        start = self.get_start_date()
        return [start + timedelta(days=i) for i in range(self.work_days)]


# Cores para cada semana do mês (usadas no CSV)
WEEK_COLORS = {
    1: "#d4edda",  # Verde claro - S1
    2: "#cce5ff",  # Azul claro - S2
    3: "#ffeeba",  # Laranja/Amarelo claro - S3
    4: "#e2d9f3",  # Roxo claro - S4
    5: "#f8d7da",  # Vermelho claro - S5
}

WEEK_COLOR_NAMES = {
    1: "Verde",
    2: "Azul", 
    3: "Laranja",
    4: "Roxo",
    5: "Vermelho",
}


@dataclass
class CalendarData:
    month: int
    year: int
    totals: Dict[int, int]
    _first_weekday: int = 0

    def last_day(self) -> int:
        _, last_day = calendar.monthrange(self.year, self.month)
        return last_day

    def to_matrix(self, first_weekday: int = 0) -> list[list[int | None]]:
        cal = calendar.Calendar(firstweekday=first_weekday)
        matrix: list[list[int | None]] = []
        for week in cal.monthdayscalendar(self.year, self.month):
            matrix.append([day if day != 0 else None for day in week])
        return matrix
    
    def to_matrix_configured(self) -> list[list[int | None]]:
        """Retorna matrix usando a configuração de primeiro dia da semana."""
        return self.to_matrix(getattr(self, '_first_weekday', 0))

    def to_matrix_with_week_configs(self, week_configs: list) -> list[list[tuple[int, int, int] | None]]:
        """
        Retorna matrix do calendário do mês atual em formato (dia, mês, ano).
        
        IMPORTANTE: O calendário SEMPRE mostra os dias normais do mês.
        As configurações de semana são usadas apenas para:
        - Definir cores/agrupamento visual
        - Cálculos de metas semanais
        
        NÃO substitui os dias do mês pelos dias da configuração.
        
        Args:
            week_configs: Lista de WeekConfig para este mês (usado para referência futura).
        
        Returns:
            Matrix de 6 semanas onde cada célula é (dia, mês, ano) ou None.
        """
        import calendar as cal_mod
        
        # Obtém o calendário base do mês
        first_weekday = getattr(self, '_first_weekday', 0)
        c = cal_mod.Calendar(firstweekday=first_weekday)
        base_matrix = c.monthdayscalendar(self.year, self.month)
        
        # Garante 6 linhas
        while len(base_matrix) < 6:
            base_matrix.append([0] * 7)
        
        matrix: list[list[tuple[int, int, int] | None]] = []
        
        for week in base_matrix:
            week_row: list[tuple[int, int, int] | None] = []
            for day in week:
                if day == 0:
                    week_row.append(None)
                else:
                    # Sempre mostra o dia do mês atual
                    week_row.append((day, self.month, self.year))
            matrix.append(week_row)
        
        return matrix

    def get_week_summary(self, day: int, targets: dict[str, WeeklyTarget]) -> dict[str, int | str]:
        """Calcula resumo semanal para um dia específico."""
        target_date = date(self.year, self.month, day)
        first_weekday = getattr(self, '_first_weekday', 0)
        
        # Usa configuração personalizada para calcular semana
        year, week = _get_custom_week_number(target_date, first_weekday)
        week_start = _get_custom_week_start(target_date, first_weekday)
        week_end = week_start + timedelta(days=6)
        
        # Calcula total feito na semana e dias restantes (sem valor)
        week_total = 0
        remaining_days = 0
        
        for i in range(7):
            current_day = week_start + timedelta(days=i)
            if (current_day.month == self.month and 
                current_day.year == self.year):
                day_value = self.totals.get(current_day.day, 0)
                week_total += day_value
                # Dia restante = dia sem valor registrado
                if day_value == 0:
                    remaining_days += 1
        
        # Busca meta configurada para a semana
        week_key = f"{year}-W{week:02d}"
        target = targets.get(week_key, WeeklyTarget(year, week, 0))
        
        expected = target.expected
        missing = max(0, expected - week_total)
        daily_needed = missing // max(1, remaining_days) if remaining_days > 0 else 0
        
        return {
            "completed": week_total,
            "expected": expected,
            "missing": missing,
            "daily_needed": daily_needed,
            "remaining_days": remaining_days
        }


class CalendarCounter:
    """Persistência de contagens por dia com suporte a meses e anos."""

    def __init__(self, data_path: Path, config_path: Path | None = None) -> None:
        """
        Inicializa os caminhos de dados e configuração.
        
        Todos os arquivos são armazenados no diretório centralizado (AppData),
        independente de onde a aplicação é executada.
        """
        # Garante que o diretório de dados existe
        app_dir = ensure_app_data_dir()
        
        # Usa o diretório centralizado para todos os arquivos internos
        self.file_path = app_dir / data_path.name
        self.config_path = app_dir / (config_path.name if config_path else "contador_config.json")
        self.targets_path = app_dir / "weekly_targets.json"
        self.week_config_path = app_dir / "weekly_config.json"
        
        # Pasta de exportação personalizada (onde o usuário quer salvar o CSV final)
        self.export_folder: Path | None = None
        
        self._data: dict[str, dict[int, int]] = {}
        self._weekly_targets: dict[str, WeeklyTarget] = {}
        self._week_configs: dict[str, WeekConfig] = {}
        self.first_weekday = 0  # Default: Monday (0=Monday, 6=Sunday)
        self._load_or_create_file()
        self._load_weekly_targets()
        self._load_week_configs()
        self._load_config()

    def _load_or_create_file(self) -> None:
        if self.file_path.exists():
            self._load_data()
        else:
            self._persist_data()

    def _load_data(self) -> None:
        with self.file_path.open("r", newline="", encoding="utf-8-sig") as csv_file:
            sample = csv_file.read(1024)
            csv_file.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ";"

            reader = csv.DictReader(csv_file, delimiter=delimiter)
            fieldnames = set(reader.fieldnames or [])
            if {"year", "month", "week"}.issubset(fieldnames) and set(WEEKDAY_LABELS).issubset(fieldnames):
                self._load_calendar_format(reader)
            elif {"year", "month"}.issubset(fieldnames):
                self._load_legacy_format(reader)

    def _load_calendar_format(self, reader: csv.DictReader) -> None:
        for row in reader:
            year = _safe_int(row.get("year"))
            month = _safe_int(row.get("month"))
            if year is None or month is None:
                continue

            key = _month_key(year, month)
            totals = self._data.setdefault(
                key,
                {day: 0 for day in range(1, calendar.monthrange(year, month)[1] + 1)},
            )

            for weekday in WEEKDAY_LABELS:
                day, value = _parse_calendar_cell(row.get(weekday))
                if day is None:
                    continue

                totals[day] = value if value is not None else totals.get(day, 0)

    def _load_legacy_format(self, reader: csv.DictReader) -> None:
        for row in reader:
            month = _safe_int(row.get("month"))
            year = _safe_int(row.get("year"))
            if month is None or year is None:
                continue

            key = _month_key(year, month)
            totals = {day: 0 for day in range(1, calendar.monthrange(year, month)[1] + 1)}
            for day_str, value in row.items():
                if day_str in {"month", "year"}:
                    continue
                day = _safe_int(day_str)
                if day is None:
                    continue
                totals[day] = _safe_int(value) or 0
            self._data[key] = totals

    def _persist_data(self) -> None:
        weekday_labels = get_weekday_labels(self.first_weekday)
        fieldnames = ["year", "month", "week", *weekday_labels]
        rows: list[dict[str, int | str]] = []

        for key, totals in sorted(self._data.items()):
            year, month = map(int, key.split("-"))
            calendar_data = CalendarData(month=month, year=year, totals=totals, _first_weekday=self.first_weekday)

            for week_index, week in enumerate(calendar_data.to_matrix_configured(), start=1):
                row: dict[str, int | str] = {name: "" for name in fieldnames}
                row["year"] = year
                row["month"] = month
                row["week"] = week_index

                for weekday_index, day in enumerate(week):
                    label = weekday_labels[weekday_index]
                    if day is None:
                        row[label] = "-"
                        continue

                    value = totals.get(day, 0)
                    row[label] = f"{day:02d} ({value})"

                rows.append(row)

        dataframe = pd.DataFrame(rows, columns=fieldnames)
        dataframe = dataframe.fillna("-")
        
        self._write_primary_file(dataframe)
        self._write_view_file()

    def _write_primary_file(self, dataframe: pd.DataFrame) -> None:
        try:
            dataframe.to_csv(self.file_path, sep=";", index=False, encoding="utf-8-sig")
        except PermissionError as error:
            # Se der erro de permissão, tenta criar arquivo temporário como fallback
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_path = self.file_path.with_name(f"{self.file_path.stem}_{timestamp}.csv")
            dataframe.to_csv(fallback_path, sep=";", index=False, encoding="utf-8-sig")
            raise PermissionError(
                (
                    f"O arquivo '{self.file_path.name}' está aberto em outro programa. "
                    f"Foi gerada uma cópia em '{fallback_path.name}'. Feche o arquivo original e tente novamente."
                )
            ) from error

    def _write_view_file(self, current_year: int | None = None, current_month: int | None = None) -> None:
        """
        Cria arquivo CSV em formato de calendário visual para Excel com células separadas.
        
        Args:
            current_year: Ano atual para filtrar (opcional). Se None, usa mês atual.
            current_month: Mês atual para filtrar (opcional). Se None, usa mês atual.
        """
        # Define período a exportar (apenas mês atual)
        if current_year is None or current_month is None:
            today = date.today()
            current_year = today.year
            current_month = today.month
        
        # Usa pasta de exportação personalizada se definida
        view_path = self.get_export_path(f"{self.file_path.stem}_view.csv")
        
        # Obtém labels configurados para os dias da semana
        weekday_labels = get_weekday_labels(self.first_weekday)
        weekday_labels_upper = [label.upper() for label in weekday_labels]
        
        # Cria estrutura de calendário visual - APENAS MÊS ATUAL
        view_rows = []
        
        key = _month_key(current_year, current_month)
        if key in self._data:
            month_totals = self._data[key]
            month_name = MONTH_NUMBER_TO_NAME.get(current_month, f"Mês {current_month}")
            calendar_data = CalendarData(month=current_month, year=current_year, totals=month_totals, _first_weekday=self.first_weekday)
            
            # Adiciona cabeçalho do mês
            month_header = {"": f"{month_name} {current_year}", "Semana": ""}
            for label in weekday_labels:
                month_header[label] = ""
            view_rows.append(month_header)
            
            # Obtém todas as configurações de semanas para este mês
            week_configs = self.get_all_week_configs_for_month(current_year, current_month)
            
            # Se há configurações personalizadas, usa elas para gerar o calendário
            if week_configs:
                # Gera calendário baseado nas configurações de semana
                for config in week_configs:
                    week_num = config.week_in_month
                    week_label = f"S{week_num} ({config.work_days}d)"
                    
                    # Obtém todas as datas desta semana configurada
                    week_dates = config.get_all_dates()
                    
                    # Calcula em qual coluna começa baseado no dia da semana do início
                    start_weekday = config.first_weekday  # 0=Seg, 6=Dom
                    # Ajusta para a ordem das colunas (baseado no first_weekday global)
                    start_col = (start_weekday - self.first_weekday) % 7
                    
                    # Se a semana não começa na primeira coluna, preenche com dias do mês anterior
                    days_before = []
                    if start_col > 0:
                        first_date = week_dates[0]
                        for i in range(start_col, 0, -1):
                            prev_date = first_date - timedelta(days=i)
                            days_before.append(prev_date)
                    
                    # Calcula quantos dias cabem na primeira linha
                    days_in_first_row = min(len(week_dates), 7 - start_col)
                    days_in_second_row = len(week_dates) - days_in_first_row
                    
                    # ===== PRIMEIRA LINHA DA SEMANA =====
                    # Linha 1: Números dos dias
                    row_days = {"": "", "Semana": week_label}
                    for label in weekday_labels:
                        row_days[label] = ""
                    
                    # Preenche dias anteriores (do mês anterior, em cinza/discreto)
                    for i, prev_date in enumerate(days_before):
                        label = weekday_labels[i]
                        month_abbr = MONTH_NUMBER_TO_NAME.get(prev_date.month, "")[:3]
                        row_days[label] = f"({prev_date.day:02d}/{month_abbr})"
                    
                    # Preenche os dias da semana configurada (primeira linha)
                    for i in range(days_in_first_row):
                        day_date = week_dates[i]
                        col_idx = start_col + i
                        label = weekday_labels[col_idx]
                        if day_date.month != current_month:
                            month_abbr = MONTH_NUMBER_TO_NAME.get(day_date.month, "")[:3]
                            row_days[label] = f"{day_date.day:02d}/{month_abbr}"
                        else:
                            row_days[label] = f"dia {day_date.day:02d}"
                    view_rows.append(row_days)
                    
                    # Linha 2: Valores de RTs (primeira linha)
                    row_rts = {"": "", "Semana": ""}
                    for label in weekday_labels:
                        row_rts[label] = ""
                    
                    # Preenche valores dos dias anteriores
                    # Se é do mês anterior, mostra o valor; se é do mesmo mês, mostra "-"
                    for i, prev_date in enumerate(days_before):
                        label = weekday_labels[i]
                        if prev_date.month != current_month:
                            # Dia do mês anterior - mostra valor
                            other_key = _month_key(prev_date.year, prev_date.month)
                            if other_key in self._data:
                                row_rts[label] = self._data[other_key].get(prev_date.day, 0)
                            else:
                                row_rts[label] = "-"
                        else:
                            # Dia do mesmo mês mas não pertence a esta semana
                            row_rts[label] = "-"
                    
                    # Preenche valores de RT da semana configurada (primeira linha)
                    for i in range(days_in_first_row):
                        day_date = week_dates[i]
                        col_idx = start_col + i
                        label = weekday_labels[col_idx]
                        if day_date.month == current_month and day_date.year == current_year:
                            row_rts[label] = month_totals.get(day_date.day, 0)
                        else:
                            other_key = _month_key(day_date.year, day_date.month)
                            if other_key in self._data:
                                row_rts[label] = self._data[other_key].get(day_date.day, 0)
                            else:
                                row_rts[label] = 0
                    view_rows.append(row_rts)
                    
                    # ===== SEGUNDA LINHA DA SEMANA (se necessário) =====
                    if days_in_second_row > 0:
                        # Linha de dias (continuação)
                        row_days2 = {"": "", "Semana": f"S{week_num} cont."}
                        for label in weekday_labels:
                            row_days2[label] = ""
                        
                        # Preenche os dias restantes (começando na coluna 0)
                        for i in range(days_in_second_row):
                            day_date = week_dates[days_in_first_row + i]
                            label = weekday_labels[i]
                            if day_date.month != current_month:
                                month_abbr = MONTH_NUMBER_TO_NAME.get(day_date.month, "")[:3]
                                row_days2[label] = f"{day_date.day:02d}/{month_abbr}"
                            else:
                                row_days2[label] = f"dia {day_date.day:02d}"
                        view_rows.append(row_days2)
                        
                        # Linha de RTs (continuação)
                        row_rts2 = {"": "", "Semana": ""}
                        for label in weekday_labels:
                            row_rts2[label] = ""
                        
                        # Preenche valores de RT restantes
                        for i in range(days_in_second_row):
                            day_date = week_dates[days_in_first_row + i]
                            label = weekday_labels[i]
                            if day_date.month == current_month and day_date.year == current_year:
                                row_rts2[label] = month_totals.get(day_date.day, 0)
                            else:
                                other_key = _month_key(day_date.year, day_date.month)
                                if other_key in self._data:
                                    row_rts2[label] = self._data[other_key].get(day_date.day, 0)
                                else:
                                    row_rts2[label] = 0
                        view_rows.append(row_rts2)
                    
                    # Linha separadora
                    separator_row = {"": "", "Semana": ""}
                    for label in weekday_labels:
                        separator_row[label] = ""
                    view_rows.append(separator_row)
            else:
                # Sem configurações personalizadas - usa calendário tradicional
                for week_num, week in enumerate(calendar_data.to_matrix_configured(), start=1):
                    week_label = f"S{week_num}"
                    
                    # Linha 1: Números dos dias
                    row_days = {"": "", "Semana": week_label}
                    for weekday_index, day in enumerate(week):
                        weekday_label = weekday_labels[weekday_index]
                        if day is None:
                            row_days[weekday_label] = ""
                        else:
                            row_days[weekday_label] = f"dia {day:02d}"
                    view_rows.append(row_days)
                    
                    # Linha 2: Valores de RTs
                    row_rts = {"": "", "Semana": ""}
                    for weekday_index, day in enumerate(week):
                        weekday_label = weekday_labels[weekday_index]
                        if day is None:
                            row_rts[weekday_label] = ""
                        else:
                            retweet_count = month_totals.get(day, 0)
                            row_rts[weekday_label] = retweet_count
                    view_rows.append(row_rts)
                    
                    # Linha separadora
                    separator_row = {"": "", "Semana": ""}
                    for label in weekday_labels:
                        separator_row[label] = ""
                    view_rows.append(separator_row)
        
        # Adiciona seção de resumo semanal (apenas mês atual)
        self._add_weekly_summary_section(view_rows, weekday_labels, current_year, current_month)
        
        if view_rows:
            columns = ["", "Semana", *weekday_labels]
            view_df = pd.DataFrame(view_rows, columns=columns)
            
            # Se o arquivo já existe, pergunta se deve sobrescrever
            if view_path.exists():
                try:
                    # Tenta sobrescrever diretamente
                    view_df.to_csv(view_path, sep=";", index=False, encoding="utf-8-sig")
                except PermissionError:
                    # Se der erro de permissão, informa que o arquivo está sendo usado
                    raise PermissionError(
                        f"O arquivo '{view_path.name}' está aberto em outro programa. "
                        f"Feche o arquivo e tente novamente."
                    )
            else:
                # Se não existe, cria normalmente
                view_df.to_csv(view_path, sep=";", index=False, encoding="utf-8-sig")
        
        # Também cria arquivo XLSX com cores (mesmo mês)
        self._write_xlsx_with_colors(current_year, current_month)

    def _write_xlsx_with_colors(self, current_year: int | None = None, current_month: int | None = None) -> None:
        """
        Cria arquivo XLSX com formatação de cores por semana.
        
        Cada semana do mês recebe uma cor diferente para facilitar
        a visualização do usuário.
        
        Args:
            current_year: Ano para filtrar (opcional). Se None, usa mês atual.
            current_month: Mês para filtrar (opcional). Se None, usa mês atual.
        """
        # Define período a exportar (apenas mês atual)
        if current_year is None or current_month is None:
            today = date.today()
            current_year = today.year
            current_month = today.month
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Se openpyxl não está instalado, ignora silenciosamente
            return
        
        # Nome do arquivo: relatorio_rts_{mes}.xlsx
        month_name = MONTH_NUMBER_TO_NAME.get(current_month, f"mes{current_month}").lower()
        xlsx_path = self.get_export_path(f"relatorio_rts_{month_name}.xlsx")
        
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Contador RTs"
            
            # Estilos
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            day_font = Font(size=9, color="666666")
            rt_font = Font(bold=True, size=16)
            week_label_font = Font(bold=True, size=10)
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Cores por semana (sem o # inicial para openpyxl)
            week_fills = {
                1: PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid"),  # Verde
                2: PatternFill(start_color="cce5ff", end_color="cce5ff", fill_type="solid"),  # Azul
                3: PatternFill(start_color="ffeeba", end_color="ffeeba", fill_type="solid"),  # Laranja
                4: PatternFill(start_color="e2d9f3", end_color="e2d9f3", fill_type="solid"),  # Roxo
                5: PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid"),  # Vermelho
            }
            
            weekday_labels = get_weekday_labels(self.first_weekday)
            current_row = 1
            
            # Processa apenas o mês atual
            key = _month_key(current_year, current_month)
            if key in self._data:
                month_totals = self._data[key]
                month_name = MONTH_NUMBER_TO_NAME.get(current_month, f"Mês {current_month}")
                calendar_data = CalendarData(
                    month=current_month, 
                    year=current_year, 
                    totals=month_totals, 
                    _first_weekday=self.first_weekday
                )
                
                # Cabeçalho do mês (mescla células)
                sheet.merge_cells(
                    start_row=current_row, 
                    start_column=1, 
                    end_row=current_row, 
                    end_column=len(weekday_labels) + 2
                )
                month_cell = sheet.cell(row=current_row, column=1, value=f"{month_name} {current_year}")
                month_cell.font = Font(bold=True, size=16, color="28a745")
                month_cell.alignment = Alignment(horizontal='center')
                current_row += 1
                
                # Cabeçalho dos dias da semana
                sheet.cell(row=current_row, column=1, value="").border = thin_border
                
                week_header = sheet.cell(row=current_row, column=2, value="Semana")
                week_header.font = Font(bold=True, size=10, color="FFFFFF")
                week_header.fill = header_fill
                week_header.alignment = Alignment(horizontal='center')
                week_header.border = thin_border
                
                for col_idx, label in enumerate(weekday_labels, start=3):
                    cell = sheet.cell(row=current_row, column=col_idx, value=label.upper())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                current_row += 1
                
                # Obtém todas as configurações de semanas para este mês
                week_configs = self.get_all_week_configs_for_month(current_year, current_month)
                
                # Se há configurações personalizadas, usa elas
                if week_configs:
                    for config in week_configs:
                        week_num = config.week_in_month
                        week_fill = week_fills.get(week_num, week_fills[1])
                        week_label = f"S{week_num} ({config.work_days}d)"
                        
                        # Obtém todas as datas desta semana configurada
                        week_dates = config.get_all_dates()
                        
                        # Calcula em qual coluna começa baseado no dia da semana do início
                        start_weekday = config.first_weekday
                        start_col = (start_weekday - self.first_weekday) % 7
                        
                        # Se a semana não começa na primeira coluna, calcula dias anteriores
                        days_before = []
                        if start_col > 0:
                            first_date = week_dates[0]
                            for i in range(start_col, 0, -1):
                                prev_date = first_date - timedelta(days=i)
                                days_before.append(prev_date)
                        
                        # Calcula quantos dias cabem na primeira linha
                        days_in_first_row = min(len(week_dates), 7 - start_col)
                        days_in_second_row = len(week_dates) - days_in_first_row
                        
                        # ===== PRIMEIRA LINHA DA SEMANA =====
                        # Linha dos dias
                        sheet.cell(row=current_row, column=1, value="").fill = week_fill
                        
                        week_cell = sheet.cell(row=current_row, column=2, value=week_label)
                        week_cell.font = week_label_font
                        week_cell.fill = week_fill
                        week_cell.alignment = Alignment(horizontal='center', vertical='center')
                        week_cell.border = thin_border
                        
                        # Preenche todas as colunas com a cor da semana
                        for col_idx in range(3, 3 + len(weekday_labels)):
                            cell = sheet.cell(row=current_row, column=col_idx)
                            cell.fill = week_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                        
                        # Preenche dias anteriores (do mês anterior) - com cor cinza
                        for i, prev_date in enumerate(days_before):
                            col_idx = 3 + i
                            cell = sheet.cell(row=current_row, column=col_idx)
                            month_abbr = MONTH_NUMBER_TO_NAME.get(prev_date.month, "")[:3]
                            cell.value = f"({prev_date.day:02d}/{month_abbr})"
                            cell.font = Font(size=8, color="888888")
                        
                        # Preenche os dias da semana configurada (primeira linha)
                        for i in range(days_in_first_row):
                            day_date = week_dates[i]
                            col_idx = 3 + start_col + i
                            cell = sheet.cell(row=current_row, column=col_idx)
                            if day_date.month != current_month:
                                month_abbr = MONTH_NUMBER_TO_NAME.get(day_date.month, "")[:3]
                                cell.value = f"{day_date.day:02d}/{month_abbr}"
                            else:
                                cell.value = f"dia {day_date.day:02d}"
                            cell.font = day_font
                        current_row += 1
                        
                        # Linha dos RTs (primeira linha)
                        sheet.cell(row=current_row, column=1, value="").fill = week_fill
                        sheet.cell(row=current_row, column=2, value="").fill = week_fill
                        
                        # Preenche todas as colunas com a cor da semana
                        for col_idx in range(3, 3 + len(weekday_labels)):
                            cell = sheet.cell(row=current_row, column=col_idx)
                            cell.fill = week_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                        
                        # Preenche valores dos dias anteriores - com cor cinza
                        # Se é do mês anterior, mostra o valor; se é do mesmo mês, mostra "-"
                        for i, prev_date in enumerate(days_before):
                            col_idx = 3 + i
                            cell = sheet.cell(row=current_row, column=col_idx)
                            if prev_date.month != current_month:
                                # Dia do mês anterior - mostra valor
                                other_key = _month_key(prev_date.year, prev_date.month)
                                if other_key in self._data:
                                    cell.value = self._data[other_key].get(prev_date.day, 0)
                                else:
                                    cell.value = "-"
                            else:
                                # Dia do mesmo mês mas não pertence a esta semana
                                cell.value = "-"
                            cell.font = Font(size=12, color="888888")
                        
                        # Preenche os valores de RT da semana configurada (primeira linha)
                        for i in range(days_in_first_row):
                            day_date = week_dates[i]
                            col_idx = 3 + start_col + i
                            cell = sheet.cell(row=current_row, column=col_idx)
                            if day_date.month == current_month and day_date.year == current_year:
                                rt_count = month_totals.get(day_date.day, 0)
                            else:
                                other_key = _month_key(day_date.year, day_date.month)
                                if other_key in self._data:
                                    rt_count = self._data[other_key].get(day_date.day, 0)
                                else:
                                    rt_count = 0
                            cell.value = rt_count
                            cell.font = rt_font
                        current_row += 1
                        
                        # ===== SEGUNDA LINHA DA SEMANA (se necessário) =====
                        if days_in_second_row > 0:
                            # Linha dos dias (continuação)
                            sheet.cell(row=current_row, column=1, value="").fill = week_fill
                            
                            cont_cell = sheet.cell(row=current_row, column=2, value=f"S{week_num} cont.")
                            cont_cell.font = Font(size=8, italic=True)
                            cont_cell.fill = week_fill
                            cont_cell.alignment = Alignment(horizontal='center', vertical='center')
                            cont_cell.border = thin_border
                            
                            # Preenche todas as colunas com a cor da semana
                            for col_idx in range(3, 3 + len(weekday_labels)):
                                cell = sheet.cell(row=current_row, column=col_idx)
                                cell.fill = week_fill
                                cell.alignment = Alignment(horizontal='center')
                                cell.border = thin_border
                            
                            # Preenche os dias restantes (começando na coluna 0 = domingo)
                            for i in range(days_in_second_row):
                                day_date = week_dates[days_in_first_row + i]
                                col_idx = 3 + i  # Começa na coluna 0
                                cell = sheet.cell(row=current_row, column=col_idx)
                                if day_date.month != current_month:
                                    month_abbr = MONTH_NUMBER_TO_NAME.get(day_date.month, "")[:3]
                                    cell.value = f"{day_date.day:02d}/{month_abbr}"
                                else:
                                    cell.value = f"dia {day_date.day:02d}"
                                cell.font = day_font
                            current_row += 1
                            
                            # Linha dos RTs (continuação)
                            sheet.cell(row=current_row, column=1, value="").fill = week_fill
                            sheet.cell(row=current_row, column=2, value="").fill = week_fill
                            
                            # Preenche todas as colunas com a cor da semana
                            for col_idx in range(3, 3 + len(weekday_labels)):
                                cell = sheet.cell(row=current_row, column=col_idx)
                                cell.fill = week_fill
                                cell.alignment = Alignment(horizontal='center')
                                cell.border = thin_border
                            
                            # Preenche os valores de RT restantes
                            for i in range(days_in_second_row):
                                day_date = week_dates[days_in_first_row + i]
                                col_idx = 3 + i
                                cell = sheet.cell(row=current_row, column=col_idx)
                                if day_date.month == current_month and day_date.year == current_year:
                                    rt_count = month_totals.get(day_date.day, 0)
                                else:
                                    other_key = _month_key(day_date.year, day_date.month)
                                    if other_key in self._data:
                                        rt_count = self._data[other_key].get(day_date.day, 0)
                                    else:
                                        rt_count = 0
                                cell.value = rt_count
                                cell.font = rt_font
                            current_row += 1
                        
                        # Linha separadora
                        current_row += 1
                else:
                    # Sem configurações personalizadas - usa calendário tradicional
                    for week_num, week in enumerate(calendar_data.to_matrix_configured(), start=1):
                        week_fill = week_fills.get(week_num, week_fills[1])
                        week_label = f"S{week_num}"
                        
                        # Linha dos dias
                        sheet.cell(row=current_row, column=1, value="").fill = week_fill
                        
                        week_cell = sheet.cell(row=current_row, column=2, value=week_label)
                        week_cell.font = week_label_font
                        week_cell.fill = week_fill
                        week_cell.alignment = Alignment(horizontal='center', vertical='center')
                        week_cell.border = thin_border
                        
                        for col_idx, day in enumerate(week, start=3):
                            cell = sheet.cell(row=current_row, column=col_idx)
                            if day is not None:
                                cell.value = f"dia {day:02d}"
                                cell.font = day_font
                            cell.fill = week_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                        current_row += 1
                        
                        # Linha dos RTs
                        sheet.cell(row=current_row, column=1, value="").fill = week_fill
                        sheet.cell(row=current_row, column=2, value="").fill = week_fill
                        
                        for col_idx, day in enumerate(week, start=3):
                            cell = sheet.cell(row=current_row, column=col_idx)
                            if day is not None:
                                rt_count = month_totals.get(day, 0)
                                cell.value = rt_count
                                cell.font = rt_font
                            cell.fill = week_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                        current_row += 1
                        
                        # Linha separadora
                        current_row += 1
            
            # Adiciona resumo semanal (apenas mês atual)
            self._add_xlsx_summary_section(sheet, current_row, weekday_labels, current_year, current_month)
            
            # Ajusta largura das colunas
            # Coluna A: dobrada para caber texto de resumo semanal
            sheet.column_dimensions['A'].width = 18
            sheet.column_dimensions['B'].width = 12
            for col_idx in range(3, len(weekday_labels) + 3):
                sheet.column_dimensions[get_column_letter(col_idx)].width = 12
            
            # Salva o arquivo
            try:
                workbook.save(xlsx_path)
            except PermissionError:
                # Se o arquivo está aberto, ignora silenciosamente
                pass
                
        except Exception:
            # Em caso de erro, ignora silenciosamente (o CSV ainda será gerado)
            pass

    def _add_xlsx_summary_section(self, sheet, start_row: int, weekday_labels: list[str],
                                    current_year: int | None = None, current_month: int | None = None) -> None:
        """
        Adiciona seção de resumo semanal no arquivo XLSX.
        
        Args:
            sheet: Planilha do openpyxl.
            start_row: Linha inicial para o resumo.
            weekday_labels: Labels dos dias da semana.
            current_year: Ano para filtrar. Se None, mostra todos.
            current_month: Mês para filtrar. Se None, mostra todos.
        """
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return
        
        current_row = start_row + 1
        
        # Cabeçalho da seção
        header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        
        sheet.merge_cells(
            start_row=current_row, 
            start_column=1, 
            end_row=current_row, 
            end_column=7
        )
        title_cell = sheet.cell(row=current_row, column=1, value="RESUMO SEMANAL - METAS E PROGRESSO")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Cabeçalhos das colunas
        columns = ["Semana", "Feito", "Esperado", "Falta", "Por Dia", "Dias Rest.", "Cor"]
        for col_idx, col_name in enumerate(columns, start=1):
            cell = sheet.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Filtra dados apenas para mês/ano atual se especificado
        if current_year and current_month:
            filter_keys = [_month_key(current_year, current_month)]
        else:
            filter_keys = sorted(self._data.keys())
        
        # Dados das semanas
        for key in filter_keys:
            if key not in self._data:
                continue
            month_totals = self._data[key]
            year_val, month_val = map(int, key.split("-"))
            calendar_data = CalendarData(
                month=month_val, 
                year=year_val, 
                totals=month_totals, 
                _first_weekday=self.first_weekday
            )
            
            for week_num, week in enumerate(calendar_data.to_matrix_configured(), start=1):
                # Busca configuração da semana
                week_config = self.get_week_config(year_val, month_val, week_num)
                
                # Busca meta usando novo formato (ano-mês-semana)
                target = self.get_weekly_target(year_val, month_val, week_num)
                expected = target.expected
                
                # Calcula totais da semana usando configuração se disponível
                week_total = 0
                remaining_days = 0
                
                if week_config:
                    # Usa datas da configuração personalizada
                    for day_date in week_config.get_all_dates():
                        day_value = 0
                        if day_date.month == month_val and day_date.year == year_val:
                            day_value = month_totals.get(day_date.day, 0)
                        else:
                            # Dia de outro mês
                            other_key = _month_key(day_date.year, day_date.month)
                            if other_key in self._data:
                                day_value = self._data[other_key].get(day_date.day, 0)
                        week_total += day_value
                        # Dia restante = dia sem valor registrado
                        if day_value == 0:
                            remaining_days += 1
                else:
                    # Usa calendário padrão
                    for day in week:
                        if day is not None:
                            day_value = month_totals.get(day, 0)
                            week_total += day_value
                            # Dia restante = dia sem valor registrado
                            if day_value == 0:
                                remaining_days += 1
                
                # Mostra semana se tem dados OU se tem configuração/meta
                has_config = week_config is not None
                has_target = expected > 0
                has_data = week_total > 0
                
                if has_data or has_config or has_target:
                    missing = max(0, expected - week_total)
                    daily_needed = missing // max(1, remaining_days) if remaining_days > 0 else 0
                    
                    # Nome da semana
                    month_name = MONTH_NUMBER_TO_NAME.get(month_val, f"M{month_val}")
                    week_display = f"{month_name[:3]}/S{week_num}"
                    if week_config:
                        week_display += f" ({week_config.work_days}d)"
                    
                    # Cor da semana
                    color_name = WEEK_COLOR_NAMES.get(week_num, "")
                    
                    # Preenche linha
                    values = [
                        week_display, 
                        week_total, 
                        expected, 
                        missing,
                        daily_needed if remaining_days > 0 else "-",
                        remaining_days if remaining_days > 0 else "-",
                        color_name
                    ]
                    
                    for col_idx, value in enumerate(values, start=1):
                        cell = sheet.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Aplica cor de fundo na coluna "Cor"
                        if col_idx == 7:  # Coluna de cor
                            week_fill = PatternFill(
                                start_color=WEEK_COLORS.get(week_num, "#ffffff").replace("#", ""),
                                end_color=WEEK_COLORS.get(week_num, "#ffffff").replace("#", ""),
                                fill_type="solid"
                            )
                            cell.fill = week_fill
                    
                    current_row += 1

    def _add_weekly_summary_section(self, view_rows: list[dict], weekday_labels: list[str], 
                                      current_year: int | None = None, current_month: int | None = None) -> None:
        """
        Adiciona seção de resumo semanal com metas no final do CSV.
        
        Args:
            view_rows: Lista de linhas do CSV.
            weekday_labels: Labels dos dias da semana.
            current_year: Ano para filtrar. Se None, mostra todos.
            current_month: Mês para filtrar. Se None, mostra todos.
        """
        # Linha separadora
        separator_row = {"": "=" * 40, "Semana": ""}
        for label in weekday_labels:
            separator_row[label] = ""
        view_rows.append(separator_row)
        
        # Cabeçalho da seção
        section_header = {"": "RESUMO SEMANAL - METAS E PROGRESSO", "Semana": ""}
        for label in weekday_labels:
            section_header[label] = ""
        view_rows.append(section_header)
        
        # Cabeçalho das colunas
        column_headers = {"": "Mês/Semana", "Semana": "Cor"}
        column_values = ["Feito", "Esperado", "Falta", "Por Dia", "Dias Rest."]
        for i, label in enumerate(weekday_labels):
            if i < len(column_values):
                column_headers[label] = column_values[i]
            else:
                column_headers[label] = ""
        view_rows.append(column_headers)
        
        # Filtra dados apenas para mês/ano atual se especificado
        if current_year and current_month:
            filter_keys = [_month_key(current_year, current_month)]
        else:
            filter_keys = sorted(self._data.keys())
        
        # Adiciona dados de cada semana que tem dados
        for key in filter_keys:
            if key not in self._data:
                continue
            month_totals = self._data[key]
            year_val, month_val = map(int, key.split("-"))
            month_name = MONTH_NUMBER_TO_NAME.get(month_val, f"M{month_val}")
            calendar_data = CalendarData(
                month=month_val, 
                year=year_val, 
                totals=month_totals, 
                _first_weekday=self.first_weekday
            )
            
            for week_num, week in enumerate(calendar_data.to_matrix_configured(), start=1):
                # Busca configuração da semana
                week_config = self.get_week_config(year_val, month_val, week_num)
                
                # Busca meta usando novo formato (ano-mês-semana)
                target = self.get_weekly_target(year_val, month_val, week_num)
                expected = target.expected
                
                # Calcula totais da semana usando configuração se disponível
                week_total = 0
                remaining_days = 0
                
                if week_config:
                    # Usa datas da configuração personalizada
                    for day_date in week_config.get_all_dates():
                        day_value = 0
                        if day_date.month == month_val and day_date.year == year_val:
                            day_value = month_totals.get(day_date.day, 0)
                        else:
                            # Dia de outro mês - busca dados corretos
                            other_key = _month_key(day_date.year, day_date.month)
                            if other_key in self._data:
                                day_value = self._data[other_key].get(day_date.day, 0)
                        week_total += day_value
                        # Dia restante = dia sem valor registrado
                        if day_value == 0:
                            remaining_days += 1
                else:
                    # Usa calendário padrão
                    for day in week:
                        if day is not None:
                            day_value = month_totals.get(day, 0)
                            week_total += day_value
                            # Dia restante = dia sem valor registrado
                            if day_value == 0:
                                remaining_days += 1
                
                # Mostra semana se tem dados OU se tem configuração/meta
                has_config = week_config is not None
                has_target = expected > 0
                has_data = week_total > 0
                
                if has_data or has_config or has_target:
                    missing = max(0, expected - week_total)
                    daily_needed = missing // max(1, remaining_days) if remaining_days > 0 else 0
                    
                    # Nome da semana no formato "Mês/S1"
                    week_display = f"{month_name[:3]}/S{week_num}"
                    if week_config:
                        week_display += f" ({week_config.work_days}d)"
                    color_name = WEEK_COLOR_NAMES.get(week_num, "")
                    
                    # Cria linha de dados
                    week_row = {"": week_display, "Semana": color_name}
                    week_values = [
                        week_total, expected, missing, 
                        daily_needed if remaining_days > 0 else "-",
                        remaining_days if remaining_days > 0 else "-"
                    ]
                    for i, label in enumerate(weekday_labels):
                        if i < len(week_values):
                            week_row[label] = week_values[i]
                        else:
                            week_row[label] = ""
                    view_rows.append(week_row)

    def get_month(self, year: int, month: int) -> CalendarData:
        key = _month_key(year, month)
        if key not in self._data:
            totals = {day: 0 for day in range(1, calendar.monthrange(year, month)[1] + 1)}
            self._data[key] = totals
        calendar_data = CalendarData(month=month, year=year, totals=self._data[key])
        # Adiciona método helper para acessar configuração
        calendar_data._first_weekday = self.first_weekday
        return calendar_data

    def increment(self, target_date: date, amount: int = 1) -> int:
        return self._apply_delta(target_date, amount)

    def decrement(self, target_date: date, amount: int = 1) -> int:
        return self._apply_delta(target_date, -amount)

    def set_value(self, target_date: date, new_value: int) -> int:
        data = self.get_month(target_date.year, target_date.month)
        value = max(0, new_value)
        data.totals[target_date.day] = value
        self._persist_data()
        return value

    def _apply_delta(self, target_date: date, amount: int) -> int:
        data = self.get_month(target_date.year, target_date.month)
        current = data.totals.get(target_date.day, 0)
        new_value = max(0, current + amount)
        data.totals[target_date.day] = new_value
        self._persist_data()
        return new_value

    def save_last_period(self, year: int, month: int) -> None:
        payload = {
            "year": year,
            "month": month,
            "first_weekday": self.first_weekday,
            "export_folder": str(self.export_folder) if self.export_folder else ""
        }
        with self.config_path.open("w", encoding="utf-8") as config_file:
            json.dump(payload, config_file, indent=2, ensure_ascii=False)
    
    def set_export_folder(self, folder_path: Path | None) -> None:
        """
        Define a pasta onde o CSV final será exportado.
        
        Args:
            folder_path: Caminho da pasta de exportação ou None para usar padrão.
        """
        if folder_path and folder_path.exists() and folder_path.is_dir():
            self.export_folder = folder_path
        else:
            self.export_folder = None
    
    def get_export_path(self, filename: str) -> Path:
        """
        Retorna o caminho completo para exportação de um arquivo.
        
        Se o usuário definiu uma pasta de exportação, usa ela.
        Caso contrário, usa o diretório padrão (AppData).
        
        Args:
            filename: Nome do arquivo a ser exportado.
        
        Returns:
            Path: Caminho completo para o arquivo.
        """
        if self.export_folder:
            return self.export_folder / filename
        return self.file_path.parent / filename

    def load_last_period(self, fallback: date) -> tuple[int, int]:
        if not self.config_path.exists():
            return fallback.year, fallback.month
        try:
            with self.config_path.open("r", encoding="utf-8") as config_file:
                payload = json.load(config_file)
            year = int(payload.get("year", fallback.year))
            month = int(payload.get("month", fallback.month))
            if 1 <= month <= 12:
                return year, month
            raise ValueError
        except (ValueError, json.JSONDecodeError):
            return fallback.year, fallback.month

    def save(self) -> None:
        self._persist_data()

    def _load_weekly_targets(self) -> None:
        """Carrega metas semanais do arquivo JSON."""
        if not self.targets_path.exists():
            return
        
        try:
            with self.targets_path.open("r", encoding="utf-8") as f:
                data = json.load(f)
                
            for week_key, target_data in data.items():
                year = target_data.get("year", 0)
                month = target_data.get("month", 0)
                week_in_month = target_data.get("week_in_month", 0)
                expected = target_data.get("expected", 0)
                
                # Compatibilidade com formato antigo (ISO week)
                if month == 0 and "week" in target_data:
                    # Formato antigo - ignora (será recriado pelo usuário)
                    continue
                
                if year > 0 and month > 0 and week_in_month > 0:
                    self._weekly_targets[week_key] = WeeklyTarget(year, month, week_in_month, expected)
                    
        except (json.JSONDecodeError, ValueError, KeyError):
            pass

    def _save_weekly_targets(self) -> None:
        """Salva metas semanais no arquivo JSON."""
        data = {}
        for week_key, target in self._weekly_targets.items():
            data[week_key] = {
                "year": target.year,
                "month": target.month,
                "week_in_month": target.week_in_month,
                "expected": target.expected
            }
        
        try:
            with self.targets_path.open("w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except (OSError, ValueError):
            pass

    def set_weekly_target(self, year: int, month: int, week_in_month: int, expected_value: int) -> None:
        """
        Define meta semanal para uma semana específica do mês.
        
        Args:
            year: Ano.
            month: Mês (1-12).
            week_in_month: Número da semana no mês (1-5).
            expected_value: Valor esperado para a semana.
        """
        week_key = f"{year}-{month:02d}-S{week_in_month}"
        self._weekly_targets[week_key] = WeeklyTarget(year, month, week_in_month, max(0, expected_value))
        self._save_weekly_targets()
        self._persist_data()  # Atualiza CSV para refletir mudanças

    def get_weekly_target(self, year: int, month: int, week_in_month: int) -> WeeklyTarget:
        """
        Obtém meta semanal para uma semana específica do mês.
        
        Args:
            year: Ano.
            month: Mês (1-12).
            week_in_month: Número da semana no mês (1-5).
        
        Returns:
            WeeklyTarget com a meta configurada ou meta zerada.
        """
        week_key = f"{year}-{month:02d}-S{week_in_month}"
        return self._weekly_targets.get(week_key, WeeklyTarget(year, month, week_in_month, 0))

    def get_weekly_summary(self, year: int, month: int, week_in_month: int) -> dict[str, int | str]:
        """
        Obtém resumo semanal para uma semana específica do mês.
        
        Args:
            year: Ano.
            month: Mês (1-12).
            week_in_month: Número da semana no mês (1-5).
        
        Returns:
            Dicionário com resumo (completed, expected, missing, daily_needed, remaining_days).
        """
        calendar_data = self.get_month(year, month)
        
        # Obtém configuração da semana
        week_config = self.get_week_config(year, month, week_in_month)
        
        # Calcula total feito na semana
        week_total = 0
        remaining_days = 0
        today = date.today()
        
        if week_config:
            # Usa datas da configuração personalizada
            for day_date in week_config.get_all_dates():
                if day_date.month == month and day_date.year == year:
                    week_total += calendar_data.totals.get(day_date.day, 0)
                else:
                    # Dia de outro mês - busca dados corretos
                    other_data = self.get_month(day_date.year, day_date.month)
                    week_total += other_data.totals.get(day_date.day, 0)
                
                if day_date > today:
                    remaining_days += 1
        else:
            # Calcula baseado no calendário padrão
            matrix = calendar_data.to_matrix_configured()
            if week_in_month <= len(matrix):
                week = matrix[week_in_month - 1]
                for day in week:
                    if day is not None:
                        week_total += calendar_data.totals.get(day, 0)
                        day_date = date(year, month, day)
                        if day_date > today:
                            remaining_days += 1
        
        # Busca meta configurada
        target = self.get_weekly_target(year, month, week_in_month)
        expected = target.expected
        missing = max(0, expected - week_total)
        daily_needed = missing // max(1, remaining_days) if remaining_days > 0 else 0
        
        return {
            "completed": week_total,
            "expected": expected,
            "missing": missing,
            "daily_needed": daily_needed,
            "remaining_days": remaining_days
        }

    def _load_config(self) -> None:
        """Carrega configurações adicionais como primeiro dia da semana e pasta de exportação."""
        if not self.config_path.exists():
            return
        
        try:
            with self.config_path.open("r", encoding="utf-8") as config_file:
                payload = json.load(config_file)
            
            # 0=Monday, 1=Tuesday, ..., 6=Sunday
            self.first_weekday = int(payload.get("first_weekday", 0))
            if not 0 <= self.first_weekday <= 6:
                self.first_weekday = 0
            
            # Pasta de exportação personalizada
            export_folder_str = payload.get("export_folder", "")
            if export_folder_str:
                export_path = Path(export_folder_str)
                if export_path.exists() and export_path.is_dir():
                    self.export_folder = export_path
                
        except (ValueError, json.JSONDecodeError, KeyError):
            self.first_weekday = 0

    def _load_week_configs(self) -> None:
        """Carrega configurações personalizadas de semanas do arquivo JSON."""
        if not self.week_config_path.exists():
            return
        
        try:
            with self.week_config_path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            
            for config_key, config_data in data.items():
                week_config = WeekConfig(
                    year=config_data.get("year", 0),
                    month=config_data.get("month", 0),
                    week_in_month=config_data.get("week_in_month", 1),
                    start_date=config_data.get("start_date", ""),
                    work_days=config_data.get("work_days", 6),
                    first_weekday=config_data.get("first_weekday", 0)
                )
                if week_config.year > 0 and week_config.month > 0:
                    self._week_configs[config_key] = week_config
                    
        except (json.JSONDecodeError, ValueError, KeyError):
            pass

    def _save_week_configs(self) -> None:
        """Salva configurações de semanas no arquivo JSON."""
        data = {}
        for config_key, config in self._week_configs.items():
            data[config_key] = {
                "year": config.year,
                "month": config.month,
                "week_in_month": config.week_in_month,
                "start_date": config.start_date,
                "work_days": config.work_days,
                "first_weekday": config.first_weekday
            }
        
        try:
            with self.week_config_path.open("w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except (OSError, ValueError):
            pass

    def set_week_config(self, year: int, month: int, week_in_month: int, 
                        start_date: date, work_days: int, first_weekday: int) -> WeekConfig:
        """
        Define configuração personalizada para uma semana específica do mês.
        
        Args:
            year: Ano da semana.
            month: Mês da semana (1-12).
            week_in_month: Número da semana no mês (1-5).
            start_date: Data de início da semana.
            work_days: Quantidade de dias de trabalho.
            first_weekday: Primeiro dia da semana (0=Seg, 6=Dom).
        
        Returns:
            WeekConfig: Configuração criada.
        """
        config = WeekConfig(
            year=year,
            month=month,
            week_in_month=week_in_month,
            start_date=start_date.strftime("%Y-%m-%d"),
            work_days=work_days,
            first_weekday=first_weekday
        )
        
        self._week_configs[config.config_key()] = config
        self._save_week_configs()
        self._persist_data()  # Atualiza CSV para refletir mudanças
        
        return config

    def get_week_config(self, year: int, month: int, week_in_month: int) -> WeekConfig | None:
        """
        Obtém configuração de uma semana específica do mês.
        
        Args:
            year: Ano da semana.
            month: Mês da semana (1-12).
            week_in_month: Número da semana no mês (1-5).
        
        Returns:
            WeekConfig ou None se não houver configuração.
        """
        config_key = f"{year}-{month:02d}-S{week_in_month}"
        return self._week_configs.get(config_key)

    def get_week_config_for_date(self, target_date: date) -> WeekConfig | None:
        """
        Obtém configuração da semana que contém a data especificada.
        
        Args:
            target_date: Data para buscar a configuração.
        
        Returns:
            WeekConfig ou None se não houver configuração.
        """
        for config in self._week_configs.values():
            if config.year == target_date.year and config.month == target_date.month:
                start = config.get_start_date()
                end = config.get_end_date()
                if start <= target_date <= end:
                    return config
        return None

    def get_all_week_configs_for_month(self, year: int, month: int) -> list[WeekConfig]:
        """
        Obtém todas as configurações de semanas para um mês.
        
        Args:
            year: Ano.
            month: Mês (1-12).
        
        Returns:
            Lista de WeekConfig ordenada por semana.
        """
        configs = []
        for config in self._week_configs.values():
            if config.year == year and config.month == month:
                configs.append(config)
        
        return sorted(configs, key=lambda c: c.week_in_month)

    def delete_week_config(self, year: int, month: int, week_in_month: int) -> bool:
        """
        Remove configuração de uma semana.
        
        Args:
            year: Ano da semana.
            month: Mês da semana (1-12).
            week_in_month: Número da semana no mês (1-5).
        
        Returns:
            True se removeu, False se não existia.
        """
        config_key = f"{year}-{month:02d}-S{week_in_month}"
        if config_key in self._week_configs:
            del self._week_configs[config_key]
            self._save_week_configs()
            self._persist_data()
            return True
        return False

    def get_week_number_in_month(self, target_date: date) -> int:
        """
        Calcula o número da semana dentro do mês (1-5).
        
        Args:
            target_date: Data de referência.
        
        Returns:
            Número da semana no mês (1-5).
        """
        # Primeiro, verifica se há configuração personalizada para esta data
        config = self.get_week_config_for_date(target_date)
        if config:
            return config.week_in_month
        
        # Caso contrário, calcula baseado no calendário padrão
        first_day = date(target_date.year, target_date.month, 1)
        first_weekday_of_month = first_day.weekday()
        
        # Ajusta para o primeiro dia da semana configurado
        adjusted_day = target_date.day + ((first_weekday_of_month - self.first_weekday) % 7)
        week_number = (adjusted_day - 1) // 7 + 1
        
        return min(week_number, 5)  # Máximo de 5 semanas por mês

    def get_week_color(self, week_in_month: int) -> str:
        """
        Retorna a cor associada a uma semana do mês.
        
        Args:
            week_in_month: Número da semana (1-5).
        
        Returns:
            Código de cor hexadecimal.
        """
        return WEEK_COLORS.get(week_in_month, "#ffffff")


# ---
# Nota de Transparência e Responsabilidade
#
# Descrição: Este arquivo contém seções de código que foram geradas
#            ou assistidas por IA.
#
# Auditoria: Todo o código foi revisado, testado e validado por
#            uma desenvolvedora humana.
#
# Tag:       @ai_generated
# Dev:       Maisa Pires
# ---
